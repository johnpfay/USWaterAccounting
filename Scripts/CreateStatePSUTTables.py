'''
Create State PSUT Tables

This script compiles 2000, 2005, and 2010 data for a given state
into a set of three PSUT tables (one for each year). The state to
process is set in line 26, and the output is stored as an Excel
file in the Data/StateData folder as "xx_PSUTt.xlsx", where "xx"
is the state abbreviation.

Original data are from the USGS state water programs.

Workflow:
    Construct the url and retrieve the data into a pandas data frame
    Melt/gather the usage columns into row values under the column name 'Group'
    Remove rows with no usage data (identified by not having "Mgal" in the 'Group' name)
    Join the remap table, which lists the destination row/column values in the PSUT

-Data Requirements:
    PSUT Template data file
    State PSUT remap table

-Package Requirements:
    pandas and openpyxl: Both can be easily installed using `pip install`

September 2017
John.Fay@Duke.edu
'''

#Specify the state and year to process
state = 'or' 

#Import modules
import sys, os, urllib
from shutil import copyfile
import pandas as pd
from openpyxl import load_workbook

#Get the input filenames
templateFilename = '../Data/Templates/StatePSUT_Template.xlsx'
remapFilename = '../Data/RemapTables/StatePSUTLookup.csv'

#Get the remap table and load as a dataframe
dfRemap = pd.read_csv(remapFilename,dtype='str',index_col="Index")

#Set the output filename
stateDataFolder = '../Data/StateData'
if os.path.exists(stateDataFolder) == False:
    os.mkdir(stateDataFolder)
outFilename = stateDataFolder + os.sep + '{0}_PSUT.xlsx'.format(state.upper())

#Copy the template to the output filename
copyfile(src=templateFilename,dst=outFilename)


#Loop through each sample year and update the PSUT with the year's data
for year in (2000,2005,2010):
    #Status
    print("Processing year {}".format(year))

    #Set the data URL path and parameters and construct the url
    print("...downloading data")
    path = 'https://waterdata.usgs.gov/{}/nwis/water_use?'.format(state)
    values = {'format':'rdb',
             'rdb_compression':'value',
             'wu_area':'County',
             'wu_year': year,
             'wu_county':'ALL',
             'wu_county_nms':'--ALL+Counties--',
             'wu_category_nms':'--ALL+Categories--'
            }
    url = path + urllib.urlencode(values)

    #Pull data in using the URL and remove the 2nd row of headers
    dfRaw = pd.read_table(url,comment='#',header=[0,1],na_values='-')
    dfRaw.columns = dfRaw.columns.droplevel(level=1)

    #Tidy the data: transform so data in each usage column become row values with a new column listing the usage type
    print("...tidying data")
    rowHeadings = ['county_cd', 'county_nm', 'state_cd', 'state_name', 'year']
    dfTidy = pd.melt(dfRaw,id_vars=rowHeadings,value_name='MGal',var_name='Group')

    #Remove rows that don't have volume data (i.e. keep only columns with 'Mgal' in the name)
    dfTidy = dfTidy[dfTidy['Group'].str.contains('Mgal')].copy(deep=True)

    #Change the type of the MGal column to float 
    dfTidy['MGal'] = dfTidy.MGal.astype('float')

    #Join the remap table
    print("...linking with remap table")
    dfAll = pd.merge(dfTidy,dfRemap,how='inner',left_on="Group",right_on="Group")

    #Open the spreadsheet template; copy the template worksheet to a new worksheet
    wb = load_workbook(filename=outFilename)
    template_ws = wb['template']
    ws = wb.copy_worksheet(template_ws)
    ws.title = str(year)

    #Loop through the three sets of row/columns and udpate values into the Excel spreadsheet
    print("...updating values in output Excel file")
    for i in (1,2,3):
        #Compute sums for column/row combinations
        dfCellData = dfAll.groupby(['Column{}'.format(i),'Row{}'.format(i)])['MGal'].sum()
        
        #Replace missing data with "n/a"
        dfCellData.fillna(value="n/a",inplace=True)

        #Iterate through all column/row pairs and update the table
        for (row,column), value in dfCellData.iteritems():
            #Set the value in the workbook
            rv = str(row)+ str(column)
            ws[rv] = value

    #Save the workbook
    wb.save(outFilename)

#Remove the template worksheet and update the output
wb.remove_sheet(wb.get_sheet_by_name('template'))
wb.save(outFilename)

